import pandas as pd
import openpyxl
import json

def get_field_names (file, sheet):
    if sheet == "": sheet = 1
    dt = pd.read_excel(file, sheet_name=sheet)
    field_names = list(dt.columns)
    #print(field_names)
    return field_names

def get_test_data_arr (file, sheet, skip):
    dd=[[]]

    if (sheet!=""):
        data = pd.read_excel(file, skiprows = skip, sheet_name=sheet)
    else:
        data = pd.read_excel(file, skiprows = skip, sheet_name=0)


    ii = data.columns.size
    jj = int(data.size/data.columns.size)

    for i in range(ii):
        if i>0:
            dd.append([])
        for j in range(jj):
            #print (str(data.loc[j][i]))
            dd[i].append( str(data.loc[j][i]).replace("nan", ""))

    return dd

#e.g.[('1','2',1),('1','3',1),('3','1',1),('2','3',1),('4','3',1)]
def get_field_ref_map  (file, sheet):
    dd=[]
    if (sheet!=""):
        data = pd.read_excel(file, sheet_name=sheet)
    else:
        data = pd.read_excel(file, sheet_name=1)

    jj = data.columns.size
    ii = int((data.size)/jj)
    c=0

    for i in range(ii):

        for j in range(jj):
            if data.loc[i][j]==1:
                dd.append([])
                dd[c].append(i+1)
                dd[c].append(j)
                dd[c].append(1)
                c = c + 1
            if i == j and i!=0 and j!=0:
                dd.append([])
                dd[c].append(j)
                dd[c].append(j)
                dd[c].append(1)
                c = c + 1
    print (dd)
    return dd

def save_gen_test_data (file, sheet, data, mode):

    if sheet=="": sheet=1
    df = pd.DataFrame(data)
    if mode == "a":
        workbook = openpyxl.load_workbook(file)
        if sheet in workbook.sheetnames:
            std = workbook.get_sheet_by_name(sheet)
            workbook.remove_sheet(std)
        writer = pd.ExcelWriter(file, mode="a", engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet, startrow=0, startcol=0, header=False, index=False)
        writer.save()
        writer.close()
        #with pd.ExcelWriter(file, mode="a", engine="openpyxl") as writer:
        #    df.to_excel(writer, sheet_name=sheet, startrow=1, startcol=1, header=False, index=False)

    else:
        with pd.ExcelWriter(file) as writer:
            df.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0, header=False, index=False)

    return sheet

def save_test_data (file, sheet, data, field_names, new_cols_map, mode):

    if sheet=="": sheet=1
    new_cols = list(new_cols_map.keys())

    #reordered_fields = field_names
    reordered_fields=[]
    print (field_names)
    print (new_cols)
    for k in range(len(field_names)):
        if k<=len(new_cols):
            x = new_cols[k]-1
            print(field_names[new_cols[k]-1])
            l=field_names[x]
            #print(x, l)
            reordered_fields.append (l)
    print (reordered_fields)


    data.insert(0,reordered_fields)
    df = pd.DataFrame(data)

    if mode == "a":
        workbook = openpyxl.load_workbook(file)
        if sheet in workbook.sheetnames:
            std = workbook.get_sheet_by_name(sheet)
            workbook.remove_sheet(std)
        writer = pd.ExcelWriter(file, mode="a", engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df.to_excel(writer, sheet,startrow=0, startcol=0, header=False, index=False)
        writer.save()
        writer.close()

    else:
        with pd.ExcelWriter(file) as writer:
            df.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0, header=False, index=False)

    result = df.to_json(orient="split")
    parsed = json.loads(result)
    res = parsed
    #res = json.dumps(parsed, indent=4)
    return res

    #return parsed